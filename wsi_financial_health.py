#!/usr/bin/env python3
"""
WSI Weekly Financial Health Report
Run every Monday morning from the AR_Automation folder.
"""

import os, re, csv
from datetime import date, datetime, timedelta
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
INPUTS_DIR = BASE_DIR / "inputs"
OUTPUT     = BASE_DIR / "WSI_Financial_Health_Report.xlsx"

AR_FILE           = INPUTS_DIR / "ar_aging.csv"
INTERACTIONS_FILE = INPUTS_DIR / "customer_interactions.csv"
AP_FILE           = INPUTS_DIR / "ap_aging.csv"
BANK_FILE         = INPUTS_DIR / "bank_transactions.csv"

# ── Colors ─────────────────────────────────────────────────────────────────────
NAVY       = "1B2A4A"
LIGHT_BLUE = "D6E4F7"
MID_BLUE   = "4472C4"
WHITE      = "FFFFFF"
GREEN_BG   = "C6EFCE"
RED_BG     = "FFC7CE"
ORANGE_BG  = "FFEB9C"
YELLOW_BG  = "FFFF99"
GRAY_BG    = "F5F5F5"
GREEN_FONT = "006100"
RED_FONT   = "9C0006"

TODAY = date.today()


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def money(s):
    """Parse a dollar string like '$1,234.56' or '1234.56' to float."""
    if s is None:
        return 0.0
    s = str(s).strip().replace("$", "").replace(",", "").replace('"', "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_date(s):
    """Try multiple date formats; return date or None."""
    if not s or str(s).strip().lower() in ("", "pending", "nan"):
        return None
    s = str(s).strip()
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d", "%-m/%-d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except (ValueError, TypeError):
            pass
    return None


def week_start(d):
    """Return Monday of the week containing d."""
    return d - timedelta(days=d.weekday())


def week_bucket(due_date):
    """
    Return 0=past_due, 1=week1, 2=week2, 3=week3, 4=week4, 5=beyond
    relative to THIS week (Monday–Sunday).
    """
    if due_date is None:
        return 5
    mon = week_start(TODAY)
    sun = mon + timedelta(days=6)
    if due_date < mon:
        return 0
    for i in range(4):
        wk_start = mon + timedelta(weeks=i)
        wk_end   = wk_start + timedelta(days=6)
        if wk_start <= due_date <= wk_end:
            return i + 1
    return 5


def fuzzy_match(name_a, name_b):
    """Case-insensitive substring or token overlap match."""
    a = re.sub(r"[^a-z0-9 ]", "", name_a.lower())
    b = re.sub(r"[^a-z0-9 ]", "", name_b.lower())
    if a in b or b in a:
        return True
    tokens_a = set(a.split())
    tokens_b = set(b.split())
    overlap = tokens_a & tokens_b - {"inc", "llc", "ltd", "the", "and", "co", "of"}
    return len(overlap) >= 2


def navy_fill():
    return PatternFill("solid", fgColor=NAVY)

def blue_fill():
    return PatternFill("solid", fgColor=LIGHT_BLUE)

def mid_fill():
    return PatternFill("solid", fgColor=MID_BLUE)

def gray_fill():
    return PatternFill("solid", fgColor=GRAY_BG)

def color_fill(hex_code):
    return PatternFill("solid", fgColor=hex_code)

def hdr_font(size=11, bold=True, color=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_row_style(ws, row, fill=None, font=None, alignment=None, number_format=None, border=None):
    for cell in row:
        if fill:       cell.fill = fill
        if font:       cell.font = font
        if alignment:  cell.alignment = alignment
        if border:     cell.border = border

def money_fmt():
    return '#,##0.00'

def set_col_widths(ws, widths):
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

def write_section_header(ws, row, col, text, colspan, fill=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font  = hdr_font(10, True, WHITE)
    cell.fill  = fill or navy_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + colspan - 1)

def freeze_row4(ws):
    ws.freeze_panes = "A5"


# ─────────────────────────────────────────────────────────────────────────────
# Data loading
# ─────────────────────────────────────────────────────────────────────────────

def load_ar(path):
    """
    Returns list of dicts with keys:
    customer, invoice, date, due_date, balance, email, days_past_due
    """
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for r in reader:
            customer = r.get("Customer full name", "").strip()
            if not customer:
                continue
            due = parse_date(r.get("Due date", ""))
            dpd = (TODAY - due).days if due else 0
            rows.append({
                "customer":      customer,
                "invoice":       str(r.get("Num", "")).strip(),
                "date":          parse_date(r.get("Date", "")),
                "due_date":      due,
                "balance":       money(r.get("Open balance", 0)),
                "email":         r.get("Email", "").strip(),
                "days_past_due": dpd,
            })
    return rows


def load_interactions(path):
    """Returns list of dicts: date, customer, invoice, type, notes"""
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for r in reader:
            customer = r.get("Customer", "").strip()
            if not customer:
                continue
            rows.append({
                "date":     parse_date(r.get("Date", "")),
                "customer": customer,
                "invoice":  str(r.get("Invoice", "")).strip(),
                "type":     r.get("Type", "").strip(),
                "notes":    r.get("Notes", "").strip(),
            })
    rows.sort(key=lambda x: x["date"] or date(2000, 1, 1), reverse=True)
    return rows


def load_ap(path):
    """
    Parse QuickBooks AP Aging Detail CSV which has section-header rows.
    Returns list of dicts: vendor, invoice, date, due_date, balance, days_past_due
    """
    rows = []
    section = "CURRENT"
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for raw in reader:
            # Section header lines like "CURRENT" or "1 - 30 days past due"
            if raw and raw[0].strip() and not raw[0].strip().startswith(","):
                val = raw[0].strip()
                if val.lower().startswith("total") or val.lower().startswith("a/p"):
                    continue
                if not any(c.isdigit() for c in val[:3]):
                    section = val
                    continue
            # Skip header / blank rows
            if len(raw) < 9:
                continue
            # Data rows: col[0] blank, col[1]=date, col[3]=num, col[4]=vendor, col[5]=due, col[6]=past_due, col[8]=open_balance
            if not raw[1].strip():
                continue
            d = parse_date(raw[1].strip())
            if d is None:
                continue
            vendor  = raw[4].strip() if len(raw) > 4 else ""
            inv_num = raw[3].strip() if len(raw) > 3 else ""
            due     = parse_date(raw[5].strip()) if len(raw) > 5 else None
            balance = money(raw[8].strip()) if len(raw) > 8 else money(raw[7].strip() if len(raw) > 7 else 0)
            dpd     = (TODAY - due).days if due else 0
            if vendor and balance > 0:
                rows.append({
                    "vendor":        vendor,
                    "invoice":       inv_num,
                    "date":          d,
                    "due_date":      due,
                    "balance":       balance,
                    "days_past_due": dpd,
                    "section":       section,
                })
    return rows


def load_bank(path):
    """
    Returns:
      transactions: list of dicts (non-pending, sorted oldest-first)
      pending:      list of pending transaction dicts
    """
    transactions = []
    pending = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for r in reader:
            raw_date = r.get("Date", "").strip()
            desc     = r.get("Description", "").strip()
            tx_desc  = r.get("Transaction Description", "").strip()
            debit    = money(r.get("Debit", 0))
            credit   = money(r.get("Credit", 0))
            balance  = money(r.get("Running Balance", 0))
            rec = {
                "date":        None,
                "description": desc,
                "tx_desc":     tx_desc,
                "debit":       debit,
                "credit":      credit,
                "balance":     balance,
            }
            if raw_date.lower() == "pending":
                pending.append(rec)
            else:
                d = parse_date(raw_date)
                if d:
                    rec["date"] = d
                    transactions.append(rec)
    transactions.sort(key=lambda x: x["date"])
    return transactions, pending


# ─────────────────────────────────────────────────────────────────────────────
# Business logic
# ─────────────────────────────────────────────────────────────────────────────

def get_ar_status(ar_row, interactions):
    """
    Return (status_emoji, last_date, latest_note) for an AR row.
    Status: ⭐ Confirmed / 🟠 Expected / 🔴 At Risk / 🟡 Pending
    """
    customer = ar_row["customer"]
    invoice  = ar_row["invoice"]
    due      = ar_row["due_date"]
    dpd      = ar_row["days_past_due"]

    # Find interactions for this customer (fuzzy) — most recent first
    matches = []
    for i in interactions:
        if fuzzy_match(customer, i["customer"]):
            matches.append(i)
        elif invoice and i["invoice"] == invoice:
            matches.append(i)
    # deduplicate
    seen = set()
    unique = []
    for m in matches:
        k = (m["date"], m["notes"][:40] if m["notes"] else "")
        if k not in seen:
            seen.add(k)
            unique.append(m)

    if not unique:
        last_date = None
        latest_note = ""
    else:
        latest = unique[0]
        last_date = latest["date"]
        latest_note = latest["notes"]

    confirmed_keywords = ["confirmed payment", "paid", "check #", "ach", "wire", "electronic payment",
                          "check run", "will pay", "pay today", "paying today"]
    risk_keywords      = ["no response", "dispute", "refused", "will not pay", "cannot pay", "issue"]

    note_lower = latest_note.lower() if latest_note else ""

    # Stale confirmation: note is > 45 days old, invoice still past due → treat as At Risk
    note_age = (TODAY - last_date).days if last_date else 999
    stale = (dpd > 0 and note_age > 45)

    if not stale and any(k in note_lower for k in confirmed_keywords):
        return "⭐ Confirmed", last_date, latest_note
    if any(k in note_lower for k in risk_keywords):
        return "🔴 At Risk", last_date, latest_note
    if stale and any(k in note_lower for k in confirmed_keywords):
        return "🔴 At Risk", last_date, f"[Stale — promised payment not received]  {latest_note}"
    if dpd > 0 and not unique:
        return "🔴 At Risk", last_date, latest_note
    if dpd > 0 and unique:
        return "🟡 Pending", last_date, latest_note
    if due and (due - TODAY).days <= 7:
        return "🟠 Expected", last_date, latest_note
    return "🟡 Pending", last_date, latest_note


def detect_fixed_costs(transactions):
    """
    Detect recurring fixed costs from bank history.
    Returns list of dicts: name, typical_amount, typical_day, last_date, next_expected, frequency
    """
    patterns = [
        ("SBA Loan",       r"SBAUSDAPMT",                 "monthly",  5),
        ("Consulting Fee", r"consulting",                  "monthly",  1),
        ("Loan Payment",   r"LOAN PAY",                    "monthly",  1),
        ("Insurance",      r"INSURANCE.*working solution", "monthly", -1),   # end of month
        ("Rent",           r"(?:RENT|BKR RENT)",           "monthly", 17),
        ("Payroll",        r"6233",                        "biweekly", None),
    ]
    results = []
    for name, pattern, freq, typical_day in patterns:
        hits = []
        for t in transactions:
            desc = (t["description"] + " " + t["tx_desc"]).lower()
            if re.search(pattern, desc, re.IGNORECASE) and t["debit"] > 0:
                hits.append(t)
        if not hits:
            continue

        if name == "Payroll":
            # Only major payroll runs (>$5k)
            major = [h for h in hits if h["debit"] > 5000]
            if not major:
                continue
            amounts = [h["debit"] for h in major]
            avg_amt = sum(amounts) / len(amounts)
            last    = max(major, key=lambda x: x["date"])
            # Bi-weekly: add 14 days from last run
            next_exp = last["date"] + timedelta(days=14)
            results.append({
                "name":          "Payroll (acct 6233)",
                "typical_amount": avg_amt,
                "last_date":     last["date"],
                "next_expected": next_exp,
                "frequency":     "Bi-weekly",
                "note":          f"Avg of {len(major)} runs",
            })
        else:
            amounts = [h["debit"] for h in hits]
            avg_amt = sum(amounts) / len(amounts)
            last    = max(hits, key=lambda x: x["date"])
            # Next expected: 1 month from last occurrence
            lm = last["date"]
            if typical_day == -1:
                # Last day of next month after last occurrence
                if lm.month == 12:
                    next_exp = date(lm.year + 1, 1, 1) - timedelta(days=1)
                else:
                    next_exp = date(lm.year, lm.month + 2, 1) - timedelta(days=1)
            else:
                nm = lm.month + 1 if lm.month < 12 else 1
                ny = lm.year if lm.month < 12 else lm.year + 1
                next_exp = date(ny, nm, typical_day)
            # Roll forward only if already in the past (strictly before today)
            while next_exp < TODAY:
                ne = next_exp
                if typical_day == -1:
                    if ne.month == 12:
                        next_exp = date(ne.year + 1, 1, 1) - timedelta(days=1)
                    else:
                        next_exp = date(ne.year, ne.month + 2, 1) - timedelta(days=1)
                else:
                    nm2 = ne.month + 1 if ne.month < 12 else 1
                    ny2 = ne.year if ne.month < 12 else ne.year + 1
                    next_exp = date(ny2, nm2, typical_day)
            results.append({
                "name":           name,
                "typical_amount": avg_amt,
                "last_date":      last["date"],
                "next_expected":  next_exp,
                "frequency":      freq.capitalize(),
                "note":           f"Avg ${avg_amt:,.2f}",
            })
    return results


def get_monthly_scorecard(transactions):
    """
    Returns list of dicts per month: month_label, money_in, money_out, net, end_balance
    From January of current year through current month.
    """
    by_month = defaultdict(lambda: {"in": 0.0, "out": 0.0, "end_balance": None})
    for t in transactions:
        if not t["date"]:
            continue
        if t["date"].year != TODAY.year:
            continue
        key = (t["date"].year, t["date"].month)
        by_month[key]["in"]  += t["credit"]
        by_month[key]["out"] += t["debit"]
        by_month[key]["end_balance"] = t["balance"]  # last seen wins (sorted ascending)

    rows = []
    for month in range(1, TODAY.month + 1):
        key = (TODAY.year, month)
        data = by_month.get(key, {"in": 0.0, "out": 0.0, "end_balance": None})
        rows.append({
            "month":       date(TODAY.year, month, 1).strftime("%B %Y"),
            "money_in":    data["in"],
            "money_out":   data["out"],
            "net":         data["in"] - data["out"],
            "end_balance": data["end_balance"],
        })
    return rows


def project_weekly_balance(bank_balance, ar_rows, ap_rows, fixed_costs, interactions):
    """
    Returns dict: week_key (0-4) → projected_balance
    0 = current (bank balance already set), 1-4 = end of each week
    """
    # AR cash in by week
    ar_in = defaultdict(float)
    for ar in ar_rows:
        bkt = week_bucket(ar["due_date"])
        status, _, _ = get_ar_status(ar, interactions)
        # Only include confirmed or expected in projection
        if "At Risk" in status:
            continue
        if bkt in (0, 1, 2, 3, 4):
            ar_in[bkt] += ar["balance"]

    # AP cash out by week
    ap_out = defaultdict(float)
    for ap in ap_rows:
        bkt = week_bucket(ap["due_date"])
        if bkt in (0, 1, 2, 3, 4):
            ap_out[bkt] += ap["balance"]

    # Fixed costs out by week
    fc_out = defaultdict(float)
    mon = week_start(TODAY)
    for fc in fixed_costs:
        ne = fc["next_expected"]
        bkt = week_bucket(ne)
        if bkt in (1, 2, 3, 4):
            fc_out[bkt] += fc["typical_amount"]
        # Also project second occurrence for bi-weekly payroll
        if fc["frequency"] == "Bi-weekly":
            ne2 = ne + timedelta(days=14)
            bkt2 = week_bucket(ne2)
            if bkt2 in (1, 2, 3, 4):
                fc_out[bkt2] += fc["typical_amount"]

    balance = bank_balance
    projected = {0: balance}
    for wk in range(1, 5):
        balance = balance + ar_in[wk] - ap_out[wk] - fc_out[wk]
        projected[wk] = balance
    return projected, ar_in, ap_out, fc_out


# ─────────────────────────────────────────────────────────────────────────────
# Excel styling helpers
# ─────────────────────────────────────────────────────────────────────────────

def style_tab_header(ws, title, subtitle=""):
    """Write a 3-row header block and return next row number."""
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 16

    # Row 1: big title
    c = ws.cell(row=1, column=1, value=title)
    c.font      = Font(name="Arial", size=14, bold=True, color=WHITE)
    c.fill      = navy_fill()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    for col in range(2, 11):
        ws.cell(row=1, column=col).fill = navy_fill()

    # Row 2: subtitle / date
    sub = subtitle or f"As of {TODAY.strftime('%A, %B %d, %Y')}"
    c2 = ws.cell(row=2, column=1, value=sub)
    c2.font      = Font(name="Arial", size=10, italic=True, color=WHITE)
    c2.fill      = navy_fill()
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    for col in range(2, 11):
        ws.cell(row=2, column=col).fill = navy_fill()

    # Row 3: blank separator
    for col in range(1, 11):
        ws.cell(row=3, column=col).fill = blue_fill()
    ws.row_dimensions[3].height = 6

    return 4


def write_col_headers(ws, row, headers, fill=None, font=None):
    f = fill or navy_fill()
    ft = font or hdr_font(10, True, WHITE)
    ws.row_dimensions[row].height = 22
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill      = f
        c.font      = ft
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
    return row + 1


def write_kv(ws, row, label, value, label_fill=None, value_fmt=None):
    """Write a label/value pair across two cells."""
    lf = label_fill or blue_fill()
    lc = ws.cell(row=row, column=1, value=label)
    lc.fill      = lf
    lc.font      = body_font(10, True)
    lc.alignment = Alignment(horizontal="left", vertical="center")
    lc.border    = thin_border()

    vc = ws.cell(row=row, column=2, value=value)
    vc.font      = body_font(10)
    vc.alignment = Alignment(horizontal="right", vertical="center")
    vc.border    = thin_border()
    if value_fmt:
        vc.number_format = value_fmt
    ws.row_dimensions[row].height = 18
    return row + 1


def write_data_row(ws, row, values, fill=None, bold=False, number_cols=None, center_cols=None):
    """Write a list of values to a row with optional formatting."""
    ws.row_dimensions[row].height = 18
    for col, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=v)
        c.font      = body_font(10, bold)
        c.alignment = Alignment(vertical="center", wrap_text=False)
        c.border    = thin_border()
        if fill:
            c.fill = fill
        if number_cols and col in number_cols:
            c.number_format = money_fmt()
            c.alignment = Alignment(horizontal="right", vertical="center")
        if center_cols and col in center_cols:
            c.alignment = Alignment(horizontal="center", vertical="center")
    return row + 1


# ─────────────────────────────────────────────────────────────────────────────
# Tab 1 — Executive Summary
# ─────────────────────────────────────────────────────────────────────────────

def build_executive_summary(ws, ar_rows, ap_rows, transactions, pending,
                             interactions, fixed_costs):
    next_row = style_tab_header(ws, "WSI Financial Health Report — Executive Summary")
    set_col_widths(ws, [28, 18, 14, 14, 14, 14, 14, 14, 14, 14])

    # ── Bank Balance ──────────────────────────────────────────────────────────
    bank_balance = transactions[-1]["balance"] if transactions else 0.0
    pending_credits = sum(p["credit"] for p in pending)
    pending_debits  = sum(p["debit"]  for p in pending)

    write_section_header(ws, next_row, 1, "CURRENT CASH POSITION", 4)
    next_row += 1
    next_row = write_kv(ws, next_row, "Current Bank Balance (BancFirst)", bank_balance, value_fmt=money_fmt())
    next_row = write_kv(ws, next_row, "Pending Credits (not yet cleared)", pending_credits, value_fmt=money_fmt())
    next_row = write_kv(ws, next_row, "Pending Debits (not yet cleared)", pending_debits, value_fmt=money_fmt())
    next_row += 1

    # ── AR / AP Totals ────────────────────────────────────────────────────────
    total_ar = sum(r["balance"] for r in ar_rows)
    total_ap = sum(r["balance"] for r in ap_rows)
    ar_past_due = sum(r["balance"] for r in ar_rows if r["days_past_due"] > 0)
    ap_past_due = sum(r["balance"] for r in ap_rows if r["days_past_due"] > 0)

    write_section_header(ws, next_row, 1, "ACCOUNTS RECEIVABLE & PAYABLE", 4)
    next_row += 1
    next_row = write_kv(ws, next_row, "Total AR Outstanding", total_ar, value_fmt=money_fmt())
    next_row = write_kv(ws, next_row, "  → Past Due AR", ar_past_due, value_fmt=money_fmt())
    next_row = write_kv(ws, next_row, "Total AP Outstanding", total_ap, value_fmt=money_fmt())
    next_row = write_kv(ws, next_row, "  → Past Due AP", ap_past_due, value_fmt=money_fmt())
    next_row += 1

    # ── This Week's Cash Flow ─────────────────────────────────────────────────
    mon = week_start(TODAY)
    sun = mon + timedelta(days=6)
    week_ar_all   = sum(r["balance"] for r in ar_rows
                        if r["due_date"] and mon <= r["due_date"] <= sun)
    # Risk-adjusted: exclude At Risk invoices (consistent with 4-week forecast)
    week_ar_in = sum(r["balance"] for r in ar_rows
                     if r["due_date"] and mon <= r["due_date"] <= sun
                     and "At Risk" not in get_ar_status(r, interactions)[0])
    week_ar_risk = week_ar_all - week_ar_in
    week_ap_out = sum(r["balance"] for r in ap_rows
                      if r["due_date"] and mon <= r["due_date"] <= sun)
    week_fc_out = sum(fc["typical_amount"] for fc in fixed_costs
                      if mon <= fc["next_expected"] <= sun)
    week_net = week_ar_in - week_ap_out - week_fc_out

    write_section_header(ws, next_row, 1,
        f"THIS WEEK'S CASH FLOW  ({mon.strftime('%b %d')} – {sun.strftime('%b %d, %Y')})", 4)
    next_row += 1
    next_row = write_kv(ws, next_row, "Expected Cash In (AR due — excl. At Risk)", week_ar_in, value_fmt=money_fmt())
    if week_ar_risk > 0:
        next_row = write_kv(ws, next_row, "  → At Risk AR this week (excluded)", week_ar_risk, value_fmt=money_fmt())
    next_row = write_kv(ws, next_row, "Expected Cash Out (AP + fixed costs)", week_ap_out + week_fc_out, value_fmt=money_fmt())

    net_cell_row = next_row
    lc = ws.cell(row=next_row, column=1, value="Projected Net This Week")
    lc.fill = blue_fill(); lc.font = body_font(10, True); lc.border = thin_border()
    lc.alignment = Alignment(horizontal="left", vertical="center")
    vc = ws.cell(row=next_row, column=2, value=week_net)
    vc.number_format = money_fmt(); vc.border = thin_border()
    vc.font = body_font(10, True, GREEN_FONT if week_net >= 0 else RED_FONT)
    vc.fill = color_fill(GREEN_BG if week_net >= 0 else RED_BG)
    vc.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[next_row].height = 18
    next_row += 2

    # ── Cash Runway ───────────────────────────────────────────────────────────
    write_section_header(ws, next_row, 1, "CASH RUNWAY INDICATOR", 4)
    next_row += 1
    monthly_sc = get_monthly_scorecard(transactions)
    recent_months = [m for m in monthly_sc if m["net"] < 0]
    if recent_months:
        avg_burn = abs(sum(m["net"] for m in recent_months[-3:]) / min(3, len(recent_months)))
        weekly_burn = avg_burn / 4.33
    else:
        avg_burn = 0
        weekly_burn = 10000  # assume positive if no negative months
    runway_weeks = bank_balance / weekly_burn if weekly_burn > 0 else 999
    runway_label = f"{runway_weeks:.1f} weeks" if runway_weeks < 100 else "12+ weeks (positive trend)"
    next_row = write_kv(ws, next_row, "Avg Monthly Burn Rate (recent neg. months)", avg_burn, value_fmt=money_fmt())
    next_row = write_kv(ws, next_row, "Estimated Runway at Current Burn", runway_label)
    next_row += 1

    # ── Monthly Scorecard Summary ──────────────────────────────────────────────
    write_section_header(ws, next_row, 1, "MONTHLY SCORECARD (YTD)", 6)
    next_row += 1
    next_row = write_col_headers(ws, next_row,
        ["Month", "Money In", "Money Out", "Net Cash Flow", "End Balance", "vs Prior Month"],
        fill=mid_fill(), font=hdr_font(10, True, WHITE))

    prior_net = None
    for sc in monthly_sc:
        vs = ""
        if prior_net is not None:
            diff = sc["net"] - prior_net
            vs = diff
        fill = color_fill(GREEN_BG) if sc["net"] >= 0 else color_fill(RED_BG)
        row_vals = [sc["month"], sc["money_in"], sc["money_out"],
                    sc["net"], sc["end_balance"] or "", vs]
        ws.row_dimensions[next_row].height = 18
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=next_row, column=col, value=v)
            c.font = body_font(10)
            c.border = thin_border()
            c.fill = fill if col in (4,) else (gray_fill() if next_row % 2 == 0 else PatternFill())
            if col in (2, 3, 4, 5, 6):
                c.number_format = money_fmt()
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        prior_net = sc["net"]
        next_row += 1
    next_row += 1

    # ── Top 5 Urgent AR ───────────────────────────────────────────────────────
    write_section_header(ws, next_row, 1, "TOP 5 URGENT AR ACTIONS", 6)
    next_row += 1
    next_row = write_col_headers(ws, next_row,
        ["Customer", "Invoice #", "Balance", "Due Date", "Days Past Due", "Status"],
        fill=mid_fill(), font=hdr_font(10, True, WHITE))

    urgent_ar = sorted(ar_rows, key=lambda x: (-x["days_past_due"], -x["balance"]))[:5]
    for ar in urgent_ar:
        status, last_d, note = get_ar_status(ar, interactions)
        dpd = ar["days_past_due"]
        dpd_display = max(0, dpd)
        fill = color_fill(RED_BG) if dpd > 30 else (
               color_fill(ORANGE_BG) if dpd > 0 else color_fill(YELLOW_BG))
        vals = [ar["customer"], ar["invoice"], ar["balance"],
                ar["due_date"].strftime("%m/%d/%y") if ar["due_date"] else "",
                dpd_display, status]
        ws.row_dimensions[next_row].height = 18
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=next_row, column=col, value=v)
            c.fill = fill; c.font = body_font(10); c.border = thin_border()
            if col == 3:
                c.number_format = money_fmt()
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        next_row += 1
    next_row += 1

    # ── Top 3 Urgent AP ───────────────────────────────────────────────────────
    write_section_header(ws, next_row, 1, "TOP 3 URGENT AP ACTIONS", 5)
    next_row += 1
    next_row = write_col_headers(ws, next_row,
        ["Vendor", "Invoice #", "Balance", "Due Date", "Days Past Due"],
        fill=mid_fill(), font=hdr_font(10, True, WHITE))

    urgent_ap = sorted(ap_rows, key=lambda x: (-x["days_past_due"], -x["balance"]))[:3]
    for ap in urgent_ap:
        dpd = ap["days_past_due"]
        dpd_display = max(0, dpd)
        fill = color_fill(RED_BG) if dpd > 0 else color_fill(ORANGE_BG)
        vals = [ap["vendor"], ap["invoice"], ap["balance"],
                ap["due_date"].strftime("%m/%d/%y") if ap["due_date"] else "",
                dpd_display]
        ws.row_dimensions[next_row].height = 18
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=next_row, column=col, value=v)
            c.fill = fill; c.font = body_font(10); c.border = thin_border()
            if col == 3:
                c.number_format = money_fmt()
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        next_row += 1

    freeze_row4(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Tab 2 — 4-Week Cash Forecast
# ─────────────────────────────────────────────────────────────────────────────

def build_cash_forecast(ws, ar_rows, ap_rows, transactions, interactions, fixed_costs):
    next_row = style_tab_header(ws, "4-Week Cash Flow Forecast")
    mon = week_start(TODAY)
    wk_labels = [f"Past Due",
                 f"Wk 1\n{mon.strftime('%b %d')}–{(mon+timedelta(6)).strftime('%d')}",
                 f"Wk 2\n{(mon+timedelta(7)).strftime('%b %d')}–{(mon+timedelta(13)).strftime('%d')}",
                 f"Wk 3\n{(mon+timedelta(14)).strftime('%b %d')}–{(mon+timedelta(20)).strftime('%d')}",
                 f"Wk 4\n{(mon+timedelta(21)).strftime('%b %d')}–{(mon+timedelta(27)).strftime('%d')}",
                 "Total", "Notes"]
    set_col_widths(ws, [32, 14, 14, 14, 14, 14, 14, 30])

    COLS = [1, 2, 3, 4, 5, 6, 7, 8]
    HDRS = ["Description"] + wk_labels
    next_row = write_col_headers(ws, next_row, HDRS, fill=navy_fill(), font=hdr_font(10, True, WHITE))
    ws.row_dimensions[next_row - 1].height = 36  # taller for wrapped week labels

    bank_balance = transactions[-1]["balance"] if transactions else 0.0

    money_cols = {2, 3, 4, 5, 6, 7}

    # ── Section: Money Coming In ──────────────────────────────────────────────
    write_section_header(ws, next_row, 1, "MONEY COMING IN  (Accounts Receivable)", 8,
                         fill=color_fill("1F6B3A"))
    next_row += 1

    bucket_totals_in = defaultdict(float)  # risk-adjusted (excludes At Risk) — used for running balance
    for ar in sorted(ar_rows, key=lambda x: x["days_past_due"], reverse=True):
        bkt = week_bucket(ar["due_date"])
        if bkt > 4:
            continue
        status, last_d, note = get_ar_status(ar, interactions)
        buckets = [0.0, 0.0, 0.0, 0.0, 0.0]  # past_due, wk1-4
        buckets[bkt] = ar["balance"]
        total = ar["balance"]
        if "At Risk" not in status:
            bucket_totals_in[bkt] += ar["balance"]
        note_str = (note[:60] + "…") if len(note) > 60 else note
        row_vals = [f"{ar['customer']} #{ar['invoice']}"] + buckets + [total, f"{status}  {note_str}"]
        fill = (color_fill(GREEN_BG)  if "Confirmed" in status else
                color_fill(RED_BG)    if "At Risk"   in status else
                color_fill(ORANGE_BG) if "Expected"  in status else
                color_fill(YELLOW_BG))
        ws.row_dimensions[next_row].height = 18
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=next_row, column=col, value=v)
            c.font = body_font(9); c.border = thin_border()
            c.fill = fill
            if col in money_cols:
                c.number_format = money_fmt()
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        next_row += 1

    # AR sub-total row (risk-adjusted — At Risk invoices shown but not counted)
    ar_total = sum(bucket_totals_in.values())
    sub_vals = ["TOTAL AR EXPECTED  (excl. At Risk)"] + [bucket_totals_in[b] for b in range(5)] + [ar_total, ""]
    ws.row_dimensions[next_row].height = 20
    for col, v in enumerate(sub_vals, 1):
        c = ws.cell(row=next_row, column=col, value=v)
        c.font = body_font(10, True); c.border = thin_border()
        c.fill = color_fill("C6EFCE")
        if col in money_cols:
            c.number_format = money_fmt()
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center")
    next_row += 2

    # ── Section: AP Bills Due ─────────────────────────────────────────────────
    write_section_header(ws, next_row, 1, "AP BILLS DUE  (Accounts Payable)", 8,
                         fill=color_fill("8B1A1A"))
    next_row += 1

    bucket_totals_ap = defaultdict(float)
    for ap in sorted(ap_rows, key=lambda x: x["due_date"] or date(2099, 1, 1)):
        bkt = week_bucket(ap["due_date"])
        if bkt > 4:
            continue
        buckets = [0.0, 0.0, 0.0, 0.0, 0.0]
        buckets[bkt] = ap["balance"]
        bucket_totals_ap[bkt] += ap["balance"]
        total = ap["balance"]
        row_vals = [f"{ap['vendor']} #{ap['invoice']}"] + buckets + [total, ""]
        dpd = ap["days_past_due"]
        fill = color_fill(RED_BG) if dpd > 0 else (
               color_fill(ORANGE_BG) if (ap["due_date"] and (ap["due_date"] - TODAY).days <= 7) else
               PatternFill())
        ws.row_dimensions[next_row].height = 18
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=next_row, column=col, value=v)
            c.font = body_font(9); c.border = thin_border()
            if fill: c.fill = fill
            if col in money_cols:
                c.number_format = money_fmt()
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        next_row += 1

    ap_total = sum(bucket_totals_ap.values())
    sub_vals = ["TOTAL AP DUE"] + [bucket_totals_ap[b] for b in range(5)] + [ap_total, ""]
    ws.row_dimensions[next_row].height = 20
    for col, v in enumerate(sub_vals, 1):
        c = ws.cell(row=next_row, column=col, value=v)
        c.font = body_font(10, True); c.border = thin_border()
        c.fill = color_fill(RED_BG)
        if col in money_cols:
            c.number_format = money_fmt()
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center")
    next_row += 2

    # ── Section: Fixed Recurring Costs ────────────────────────────────────────
    write_section_header(ws, next_row, 1, "FIXED RECURRING COSTS  (Auto-detected from bank history)", 8,
                         fill=color_fill("5B3A8B"))
    next_row += 1

    bucket_totals_fc = defaultdict(float)
    for fc in fixed_costs:
        bkt = week_bucket(fc["next_expected"])
        # Fixed recurring costs: extend lookahead to 35 days — fold week-5 items
        # into week 4 so nothing is silently dropped from the running balance.
        if bkt == 5 and 0 <= (fc["next_expected"] - TODAY).days <= 34:
            bkt = 4
        buckets = [0.0, 0.0, 0.0, 0.0, 0.0]
        if bkt in range(5):
            buckets[bkt] = fc["typical_amount"]
            bucket_totals_fc[bkt] += fc["typical_amount"]
        total = fc["typical_amount"] if bkt in range(5) else 0
        ne_label = fc["next_expected"].strftime("%b %-d") if fc["next_expected"] else ""
        note = f"{fc['frequency']}  |  Last: {fc['last_date'].strftime('%m/%d/%y') if fc['last_date'] else 'N/A'}  |  Next: {ne_label}"
        row_vals = [fc["name"]] + buckets + [total, note]
        ws.row_dimensions[next_row].height = 18
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=next_row, column=col, value=v)
            c.font = body_font(10); c.border = thin_border()
            c.fill = color_fill(LIGHT_BLUE)
            if col in money_cols:
                c.number_format = money_fmt()
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        next_row += 1

        # Second payroll occurrence within 4 weeks (or 35-day overflow)
        if fc["frequency"] == "Bi-weekly":
            ne2 = fc["next_expected"] + timedelta(days=14)
            bkt2 = week_bucket(ne2)
            if bkt2 == 5 and 0 <= (ne2 - TODAY).days <= 34:
                bkt2 = 4
            if bkt2 in range(1, 5):
                buckets2 = [0.0, 0.0, 0.0, 0.0, 0.0]
                buckets2[bkt2] = fc["typical_amount"]
                bucket_totals_fc[bkt2] += fc["typical_amount"]
                row_vals2 = [fc["name"] + " (2nd)"] + buckets2 + [fc["typical_amount"], ""]
                ws.row_dimensions[next_row].height = 18
                for col, v in enumerate(row_vals2, 1):
                    c = ws.cell(row=next_row, column=col, value=v)
                    c.font = body_font(10); c.border = thin_border()
                    c.fill = color_fill(LIGHT_BLUE)
                    if col in money_cols:
                        c.number_format = money_fmt()
                        c.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        c.alignment = Alignment(horizontal="left", vertical="center")
                next_row += 1

    fc_total = sum(bucket_totals_fc.values())
    sub_vals = ["TOTAL FIXED COSTS"] + [bucket_totals_fc[b] for b in range(5)] + [fc_total, ""]
    ws.row_dimensions[next_row].height = 20
    for col, v in enumerate(sub_vals, 1):
        c = ws.cell(row=next_row, column=col, value=v)
        c.font = body_font(10, True); c.border = thin_border()
        c.fill = color_fill("E2D5F5")
        if col in money_cols:
            c.number_format = money_fmt()
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center")
    next_row += 2

    # ── Weekly Net Cash Flow row (shows why each week's balance moves) ───────────
    net_row = ["NET CASH FLOW THIS WEEK", ""]
    for wk in range(1, 5):
        net_row.append(bucket_totals_in[wk] - bucket_totals_ap[wk] - bucket_totals_fc[wk])
    net_row.append("")
    net_row.append("AR in − AP out − Fixed costs")

    ws.row_dimensions[next_row].height = 20
    for col, v in enumerate(net_row, 1):
        c = ws.cell(row=next_row, column=col, value=v)
        c.border = thin_border()
        if col in {3, 4, 5, 6} and isinstance(v, (int, float)):
            c.number_format = money_fmt()
            c.alignment = Alignment(horizontal="right", vertical="center")
            positive = v >= 0
            c.font = body_font(10, True, GREEN_FONT if positive else RED_FONT)
            c.fill = color_fill(GREEN_BG if positive else RED_BG)
        else:
            c.font = body_font(10, True)
            c.fill = color_fill(LIGHT_BLUE)
            c.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
    next_row += 1

    # ── Projected Running Balance ─────────────────────────────────────────────
    balance = bank_balance
    bal_row = ["PROJECTED RUNNING BALANCE"]
    bal_row.append("")  # past due col — no cumulative used here
    for wk in range(1, 5):
        in_  = bucket_totals_in[wk]
        out_ = bucket_totals_ap[wk] + bucket_totals_fc[wk]
        balance += in_ - out_
        bal_row.append(balance)
    bal_row.append("")  # total col
    bal_row.append(f"Starting balance: ${bank_balance:,.2f}")

    ws.row_dimensions[next_row].height = 22
    for col, v in enumerate(bal_row, 1):
        c = ws.cell(row=next_row, column=col, value=v)
        c.font = body_font(11, True, WHITE)
        c.fill = navy_fill()
        c.border = thin_border()
        if col in {3, 4, 5, 6}:
            c.number_format = money_fmt()
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center")

    freeze_row4(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Tab 3 — AR Collections Detail
# ─────────────────────────────────────────────────────────────────────────────

def build_ar_detail(ws, ar_rows, interactions):
    next_row = style_tab_header(ws, "AR Collections Detail — All Open Invoices")
    set_col_widths(ws, [35, 12, 14, 12, 14, 18, 14, 50])

    headers = ["Customer", "Invoice #", "Amount", "Due Date", "Days Past Due",
               "Status", "Last Contact", "Latest Note"]
    next_row = write_col_headers(ws, next_row, headers)

    sorted_ar = sorted(ar_rows, key=lambda x: (-x["days_past_due"], -x["balance"]))

    for ar in sorted_ar:
        status, last_d, note = get_ar_status(ar, interactions)
        dpd = ar["days_past_due"]
        dpd_display = dpd if dpd > 0 else 0

        if "Confirmed" in status:
            fill = color_fill(GREEN_BG)
        elif "At Risk" in status:
            fill = color_fill(RED_BG)
        elif "Expected" in status:
            fill = color_fill(ORANGE_BG)
        else:
            fill = color_fill(YELLOW_BG)

        vals = [
            ar["customer"],
            ar["invoice"],
            ar["balance"],
            ar["due_date"].strftime("%m/%d/%y") if ar["due_date"] else "",
            dpd_display,
            status,
            last_d.strftime("%m/%d/%y") if last_d else "No contact",
            (note[:120] + "…") if len(note) > 120 else note,
        ]
        ws.row_dimensions[next_row].height = 18
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=next_row, column=col, value=v)
            c.fill = fill; c.font = body_font(10); c.border = thin_border()
            if col == 3:
                c.number_format = money_fmt()
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif col in (4, 5, 6, 7):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        next_row += 1

    # Totals row
    total = sum(r["balance"] for r in ar_rows)
    ws.row_dimensions[next_row].height = 20
    c1 = ws.cell(row=next_row, column=1, value=f"TOTAL  ({len(ar_rows)} invoices)")
    c1.font = body_font(10, True); c1.fill = navy_fill(); c1.font = hdr_font(10, True, WHITE)
    c1.border = thin_border()
    c2 = ws.cell(row=next_row, column=3, value=total)
    c2.number_format = money_fmt()
    c2.font = hdr_font(10, True, WHITE); c2.fill = navy_fill(); c2.border = thin_border()
    c2.alignment = Alignment(horizontal="right", vertical="center")
    for col in [2, 4, 5, 6, 7, 8]:
        c = ws.cell(row=next_row, column=col)
        c.fill = navy_fill(); c.border = thin_border()

    freeze_row4(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Tab 4 — AP + Fixed Costs Detail
# ─────────────────────────────────────────────────────────────────────────────

def build_ap_detail(ws, ap_rows, fixed_costs):
    next_row = style_tab_header(ws, "AP + Fixed Costs Detail")
    set_col_widths(ws, [35, 15, 14, 12, 14, 18])

    headers = ["Vendor", "Invoice #", "Amount", "Due Date", "Days Past Due", "Priority"]
    next_row = write_col_headers(ws, next_row, headers)

    sorted_ap = sorted(ap_rows, key=lambda x: x["due_date"] or date(2099, 1, 1))

    for ap in sorted_ap:
        dpd = ap["days_past_due"]
        dpd_display = max(0, dpd)
        if dpd > 30:
            priority = "🔴 URGENT"
            fill = color_fill(RED_BG)
        elif dpd > 0:
            priority = "🟠 Past Due"
            fill = color_fill(ORANGE_BG)
        elif ap["due_date"] and (ap["due_date"] - TODAY).days <= 7:
            priority = "🟡 Due This Week"
            fill = color_fill(YELLOW_BG)
        elif ap["due_date"] and (ap["due_date"] - TODAY).days <= 14:
            priority = "🟢 Due Next Week"
            fill = color_fill(GREEN_BG)
        else:
            fill = PatternFill()
            priority = "Upcoming"

        vals = [ap["vendor"], ap["invoice"], ap["balance"],
                ap["due_date"].strftime("%m/%d/%y") if ap["due_date"] else "",
                dpd_display, priority]
        ws.row_dimensions[next_row].height = 18
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=next_row, column=col, value=v)
            if fill: c.fill = fill
            c.font = body_font(10); c.border = thin_border()
            if col == 3:
                c.number_format = money_fmt()
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif col in (4, 5, 6):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        next_row += 1

    # AP total
    ap_total = sum(r["balance"] for r in ap_rows)
    ws.row_dimensions[next_row].height = 20
    for col in range(1, 7):
        c = ws.cell(row=next_row, column=col)
        c.fill = navy_fill(); c.border = thin_border()
        c.font = hdr_font(10, True, WHITE)
    ws.cell(row=next_row, column=1).value = f"TOTAL AP  ({len(ap_rows)} bills)"
    c3 = ws.cell(row=next_row, column=3, value=ap_total)
    c3.number_format = money_fmt()
    c3.alignment = Alignment(horizontal="right", vertical="center")
    next_row += 2

    # ── Fixed Recurring Costs ─────────────────────────────────────────────────
    write_section_header(ws, next_row, 1, "FIXED RECURRING COSTS  (Auto-detected)", 6)
    next_row += 1

    fc_headers = ["Cost Name", "Frequency", "Typical Amount", "Last Paid", "Next Expected", "Note"]
    next_row = write_col_headers(ws, next_row, fc_headers, fill=mid_fill())

    for fc in fixed_costs:
        vals = [
            fc["name"],
            fc["frequency"],
            fc["typical_amount"],
            fc["last_date"].strftime("%m/%d/%y") if fc["last_date"] else "",
            fc["next_expected"].strftime("%m/%d/%y") if fc["next_expected"] else "",
            fc["note"],
        ]
        ws.row_dimensions[next_row].height = 18
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=next_row, column=col, value=v)
            c.fill = color_fill(LIGHT_BLUE); c.font = body_font(10); c.border = thin_border()
            if col == 3:
                c.number_format = money_fmt()
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif col in (4, 5):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        next_row += 1

    freeze_row4(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Tab 5 — Monthly Scorecard
# ─────────────────────────────────────────────────────────────────────────────

def build_monthly_scorecard(ws, transactions):
    next_row = style_tab_header(ws, "Monthly Scorecard — Cash Flow by Month")
    set_col_widths(ws, [18, 16, 16, 16, 16, 16])

    headers = ["Month", "Total Money In", "Total Money Out", "Net Cash Flow",
               "End Balance", "vs Prior Month"]
    next_row = write_col_headers(ws, next_row, headers)

    monthly = get_monthly_scorecard(transactions)
    prior_net = None

    for sc in monthly:
        vs = ""
        if prior_net is not None:
            vs = sc["net"] - prior_net

        net_positive = sc["net"] >= 0
        row_fill = color_fill(GREEN_BG) if net_positive else color_fill(RED_BG)

        vals = [sc["month"], sc["money_in"], sc["money_out"],
                sc["net"], sc["end_balance"] or 0, vs]

        ws.row_dimensions[next_row].height = 22
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=next_row, column=col, value=v)
            c.border = thin_border()
            c.font = body_font(10, col == 4)  # bold net column
            if col == 1:
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.number_format = money_fmt()
                c.alignment = Alignment(horizontal="right", vertical="center")
            # Color the net column; alternate gray for others
            if col == 4:
                c.fill = row_fill
                if not net_positive:
                    c.font = Font(name="Arial", size=10, bold=True, color=RED_FONT)
            elif next_row % 2 == 0:
                c.fill = gray_fill()

        prior_net = sc["net"]
        next_row += 1

    # Totals
    next_row += 1
    ws.row_dimensions[next_row].height = 22
    total_in  = sum(m["money_in"]  for m in monthly)
    total_out = sum(m["money_out"] for m in monthly)
    total_net = total_in - total_out
    end_bal   = monthly[-1]["end_balance"] if monthly else 0

    totals_vals = ["YTD TOTAL", total_in, total_out, total_net, end_bal, ""]
    for col, v in enumerate(totals_vals, 1):
        c = ws.cell(row=next_row, column=col, value=v)
        c.font = hdr_font(10, True, WHITE)
        c.fill = navy_fill(); c.border = thin_border()
        if col > 1:
            c.number_format = money_fmt()
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center")

    freeze_row4(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def validate_inputs():
    missing = []
    for f in [AR_FILE, INTERACTIONS_FILE, AP_FILE, BANK_FILE]:
        if not f.exists():
            missing.append(str(f))
    if missing:
        print("\n❌  Missing input files:")
        for m in missing:
            print(f"    {m}")
        print(f"\nPlace all 4 CSV files in:  {INPUTS_DIR}/\n")
        return False
    return True


def main():
    print("\n" + "="*60)
    print("  WSI Financial Health Report")
    print(f"  Generated: {TODAY.strftime('%A, %B %d, %Y')}")
    print("="*60)

    if not validate_inputs():
        return

    print("\nLoading data...")
    ar_rows      = load_ar(AR_FILE)
    interactions = load_interactions(INTERACTIONS_FILE)
    ap_rows      = load_ap(AP_FILE)
    transactions, pending = load_bank(BANK_FILE)
    fixed_costs  = detect_fixed_costs(transactions)

    bank_balance = transactions[-1]["balance"] if transactions else 0.0
    total_ar     = sum(r["balance"] for r in ar_rows)
    total_ap     = sum(r["balance"] for r in ap_rows)
    ar_past_due  = sum(r["balance"] for r in ar_rows if r["days_past_due"] > 0)

    print(f"\n  Bank Balance:     ${bank_balance:>12,.2f}")
    print(f"  Total AR:         ${total_ar:>12,.2f}   (${ar_past_due:,.2f} past due)")
    print(f"  Total AP:         ${total_ap:>12,.2f}")
    print(f"  AR Invoices:      {len(ar_rows)} open")
    print(f"  AP Bills:         {len(ap_rows)} open")
    print(f"  Bank transactions:{len(transactions)} records ({transactions[0]['date']} – {transactions[-1]['date']})" if transactions else "  No bank transactions.")
    print(f"  Pending items:    {len(pending)}")
    print(f"  Fixed costs det.: {len(fixed_costs)}")
    if fixed_costs:
        for fc in fixed_costs:
            print(f"    • {fc['name']:<25} ${fc['typical_amount']:>10,.2f}  next: {fc['next_expected']}")

    print("\nBuilding Excel report...")

    wb = openpyxl.Workbook()
    tabs = [
        ("Executive Summary",    lambda ws: build_executive_summary(ws, ar_rows, ap_rows, transactions, pending, interactions, fixed_costs)),
        ("4-Week Cash Forecast", lambda ws: build_cash_forecast(ws, ar_rows, ap_rows, transactions, interactions, fixed_costs)),
        ("AR Collections",       lambda ws: build_ar_detail(ws, ar_rows, interactions)),
        ("AP + Fixed Costs",     lambda ws: build_ap_detail(ws, ap_rows, fixed_costs)),
        ("Monthly Scorecard",    lambda ws: build_monthly_scorecard(ws, transactions)),
    ]

    for i, (tab_name, builder) in enumerate(tabs):
        if i == 0:
            ws = wb.active
            ws.title = tab_name
        else:
            ws = wb.create_sheet(title=tab_name)
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        builder(ws)
        print(f"  ✓ {tab_name}")

    wb.save(OUTPUT)
    print(f"\n✅  Report saved to:\n    {OUTPUT}\n")


if __name__ == "__main__":
    main()
